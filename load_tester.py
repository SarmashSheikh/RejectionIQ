import sys
import os
import time
import json
import math
import argparse
import statistics
import concurrent.futures
import urllib.request
import urllib.parse
import urllib.error
from datetime import datetime

# Default configuration
DEFAULT_HOST = "http://127.0.0.1:8000"
DEFAULT_USERS = 100
DEFAULT_DURATION = 60  # seconds

class LoadTester:
    def __init__(self, host: str, num_users: int, duration_sec: int):
        self.host = host.rstrip('/')
        self.num_users = num_users
        self.duration_sec = duration_sec
        self.results = []
        self.auth_token = None
        self.status_counts = {}
        self.start_time = 0
        self.end_time = 0
        
    def obtain_auth_token(self):
        """Obtain JWT auth token for authenticated load testing."""
        url = self.host + "/api/auth/login"
        payload = urllib.parse.urlencode({
            "username": "demo@rejectioniq.com",
            "password": "demo1234"
        }).encode('utf-8')
        
        req = urllib.request.Request(url, data=payload, method="POST")
        req.add_header("Content-Type", "application/x-www-form-urlencoded")
        
        try:
            with urllib.request.urlopen(req, timeout=5) as resp:
                data = json.loads(resp.read().decode('utf-8'))
                self.auth_token = data.get("access_token")
                print(f"[+] Successfully authenticated load test user 'demo@rejectioniq.com'. Token acquired.")
        except Exception as e:
            print(f"[!] Warning: Could not authenticate load test user: {e}. Running in unauthenticated mode.")
            self.auth_token = None

    def get_test_endpoints(self):
        return [
            {"name": "Root Welcome", "path": "/", "method": "GET", "auth": False},
            {"name": "Health Check", "path": "/api/health", "method": "GET", "auth": False},
            {"name": "User Profile", "path": "/api/users/me", "method": "GET", "auth": True},
            {"name": "Rejections List", "path": "/api/rejections/", "method": "GET", "auth": True},
            {"name": "Pattern Analysis", "path": "/api/analysis/pattern", "method": "GET", "auth": True},
        ]

    def send_request(self, endpoint_info):
        url = self.host + endpoint_info["path"]
        method = endpoint_info["method"]
        name = endpoint_info["name"]
        
        req = urllib.request.Request(url, method=method)
        req.add_header("User-Agent", "RejectionIQ-LoadTester/1.0")
        
        if endpoint_info.get("auth") and self.auth_token:
            req.add_header("Authorization", f"Bearer {self.auth_token}")
            
        t0 = time.perf_counter()
        status_code = 0
        success = False
        error_msg = ""
        
        try:
            with urllib.request.urlopen(req, timeout=10) as resp:
                status_code = resp.getcode()
                _ = resp.read()
                success = (200 <= status_code < 300)
        except urllib.error.HTTPError as e:
            status_code = e.code
            error_msg = str(e.reason)
            success = False
        except urllib.error.URLError as e:
            status_code = 0
            error_msg = str(e.reason)
            success = False
        except Exception as e:
            status_code = 0
            error_msg = str(e)
            success = False
            
        t1 = time.perf_counter()
        latency_ms = (t1 - t0) * 1000.0
        
        return {
            "name": name,
            "path": endpoint_info["path"],
            "status_code": status_code,
            "latency_ms": latency_ms,
            "success": success,
            "error": error_msg,
            "timestamp": t1
        }

    def worker_loop(self, worker_id, stop_time, endpoints):
        worker_results = []
        ep_index = worker_id % len(endpoints)
        
        while time.time() < stop_time:
            ep = endpoints[ep_index]
            res = self.send_request(ep)
            worker_results.append(res)
            ep_index = (ep_index + 1) % len(endpoints)
            time.sleep(0.001)
            
        return worker_results

    def run(self):
        print("=" * 75)
        print("                REJECTIONIQ BASELINE / LOAD TESTING               ")
        print("=" * 75)
        print(f" Target Server       : {self.host}")
        print(f" Concurrent Users    : {self.num_users} Virtual Users")
        print(f" Test Duration       : {self.duration_sec} Seconds continuous")
        print("=" * 75)

        self.obtain_auth_token()
        endpoints = self.get_test_endpoints()
        self.endpoint_stats = {ep["name"]: [] for ep in endpoints}

        print("\n[+] Initializing 100 worker pool and beginning continuous load test...\n")

        self.start_time = time.time()
        stop_time = self.start_time + self.duration_sec
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=self.num_users) as executor:
            futures = [
                executor.submit(self.worker_loop, i, stop_time, endpoints)
                for i in range(self.num_users)
            ]
            
            while time.time() < stop_time:
                elapsed = time.time() - self.start_time
                pct = min(100.0, (elapsed / self.duration_sec) * 100.0)
                remaining = max(0.0, self.duration_sec - elapsed)
                sys.stdout.write(f"\rProgress: [{pct:5.1f}%] | Elapsed: {elapsed:4.1f}s / {self.duration_sec}s | Remaining: {remaining:4.1f}s | Active VUs: {self.num_users}")
                sys.stdout.flush()
                time.sleep(0.5)
                
            print("\n\n[+] Test execution complete. Aggregating performance metrics...\n")
            
            all_worker_results = []
            for f in concurrent.futures.as_completed(futures):
                try:
                    all_worker_results.extend(f.result())
                except Exception as e:
                    print(f"Worker exception: {e}")
                    
        self.end_time = time.time()
        self.results = all_worker_results
        return self.analyze_and_report()

    def analyze_and_report(self):
        total_requests = len(self.results)
        actual_duration = self.end_time - self.start_time
        rps = total_requests / actual_duration if actual_duration > 0 else 0
        
        if total_requests == 0:
            print("[-] Error: No requests were recorded. Is the backend server running?")
            return None
            
        latencies = [r["latency_ms"] for r in self.results]
        latencies.sort()
        
        successful_reqs = [r for r in self.results if r["success"]]
        failed_reqs = [r for r in self.results if not r["success"]]
        
        success_rate = (len(successful_reqs) / total_requests) * 100.0
        failure_rate = (len(failed_reqs) / total_requests) * 100.0
        
        avg_latency = statistics.mean(latencies)
        min_latency = min(latencies)
        max_latency = max(latencies)
        median_latency = statistics.median(latencies)
        
        def percentile(N, percent):
            if not N:
                return 0
            k = (len(N) - 1) * (percent / 100.0)
            f = math.floor(k)
            c = math.ceil(k)
            if f == c:
                return N[int(k)]
            d0 = N[int(f)] * (c - k)
            d1 = N[int(c)] * (k - f)
            return d0 + d1

        p90_latency = percentile(latencies, 90)
        p95_latency = percentile(latencies, 95)
        p99_latency = percentile(latencies, 99)
        
        for r in self.results:
            sc = r["status_code"]
            self.status_counts[sc] = self.status_counts.get(sc, 0) + 1
            if r["name"] in self.endpoint_stats:
                self.endpoint_stats[r["name"]].append(r["latency_ms"])
            
        report = {
            "test_info": {
                "target_host": self.host,
                "virtual_users": self.num_users,
                "target_duration_sec": self.duration_sec,
                "actual_duration_sec": round(actual_duration, 2),
                "timestamp": datetime.now().isoformat()
            },
            "throughput": {
                "total_requests": total_requests,
                "successful_requests": len(successful_reqs),
                "failed_requests": len(failed_reqs),
                "requests_per_second": round(rps, 2),
                "success_rate_pct": round(success_rate, 2),
                "failure_rate_pct": round(failure_rate, 2)
            },
            "response_time_ms": {
                "avg": round(avg_latency, 2),
                "min": round(min_latency, 2),
                "max": round(max_latency, 2),
                "median_p50": round(median_latency, 2),
                "p90": round(p90_latency, 2),
                "p95": round(p95_latency, 2),
                "p99": round(p99_latency, 2)
            },
            "status_codes": self.status_counts,
            "endpoint_breakdown": {}
        }
        
        for name, l_list in self.endpoint_stats.items():
            if l_list:
                report["endpoint_breakdown"][name] = {
                    "count": len(l_list),
                    "rps": round(len(l_list) / actual_duration, 2),
                    "min_ms": round(min(l_list), 2),
                    "avg_ms": round(statistics.mean(l_list), 2),
                    "max_ms": round(max(l_list), 2)
                }

        # Print Executive Terminal Summary
        print("=" * 75)
        print("                        LOAD TEST EXECUTION RESULTS                       ")
        print("=" * 75)
        print(f" Total Requests Sent : {total_requests:,} requests")
        print(f" Actual Duration     : {actual_duration:.2f} seconds")
        print(f" Requests Per Second : {rps:.2f} req/sec  <--- (RPS)")
        print(f" Success Rate        : {success_rate:.2f}% ({len(successful_reqs)} ok, {len(failed_reqs)} failed)")
        print("-" * 75)
        print(" RESPONSE TIME BREAKDOWN (Latency in ms)")
        print("-" * 75)
        print(f"  • Average Latency  : {avg_latency:.2f} ms  (Target: ~250ms)")
        print(f"  • Min Latency      : {min_latency:.2f} ms   (Target: ~50ms)")
        print(f"  • Max Latency      : {max_latency:.2f} ms (Target: ~1500ms)")
        print(f"  • Median (p50)     : {median_latency:.2f} ms")
        print(f"  • 90th Percentile  : {p90_latency:.2f} ms")
        print(f"  • 95th Percentile  : {p95_latency:.2f} ms")
        print(f"  • 99th Percentile  : {p99_latency:.2f} ms")
        print("-" * 75)
        print(" ENDPOINT BREAKDOWN")
        print("-" * 75)
        print(f" {'Endpoint Name':<22} | {'Reqs':<7} | {'RPS':<8} | {'Min (ms)':<8} | {'Avg (ms)':<8} | {'Max (ms)':<8}")
        print(f" {'-'*22}-+-{'-'*7}-+-{'-'*8}-+-{'-'*8}-+-{'-'*8}-+-{'-'*8}")
        for name, stats in report["endpoint_breakdown"].items():
            print(f" {name:<22} | {stats['count']:<7} | {stats['rps']:<8.1f} | {stats['min_ms']:<8.1f} | {stats['avg_ms']:<8.1f} | {stats['max_ms']:<8.1f}")
        print("=" * 75)
        print(f" HTTP Status Codes   : {self.status_counts}")
        print("=" * 75)
        
        json_filename = "load_test_results.json"
        with open(json_filename, "w") as f:
            json.dump(report, f, indent=2)
        print(f"\n[+] Raw results saved to: {os.path.abspath(json_filename)}")

        md_filename = "Load_Test_Report.md"
        self.generate_markdown_report(report, md_filename)
        print(f"[+] Formatted Markdown report saved to: {os.path.abspath(md_filename)}")

        excel_filename = "Load_Test_Report_RejectionIQ.xlsx"
        self.generate_excel_report(report, excel_filename)
        print(f"[+] Formatted Excel report saved to: {os.path.abspath(excel_filename)}\n")

        return report

    def generate_excel_report(self, report, filename):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Load Test Summary"

            title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            bold_font = Font(name="Calibri", size=11, bold=True)
            regular_font = Font(name="Calibri", size=11)

            navy_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            soft_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )

            # Title
            ws.merge_cells("A1:F1")
            ws["A1"] = "RejectionIQ - Baseline & Load Testing Executive Report"
            ws["A1"].font = title_font
            ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

            ws["A2"] = f"Execution Timestamp: {report['test_info']['timestamp']} | Target: {report['test_info']['target_host']}"
            ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="595959")

            # Summary Table
            ws.append([])
            ws.append(["Parameter", "Value", "Description"])
            row_idx = 4
            for cell in ws[row_idx]:
                cell.font = header_font
                cell.fill = navy_fill
                cell.alignment = Alignment(horizontal="left", vertical="center")

            summary_data = [
                ("Virtual Users (VUs)", report['test_info']['virtual_users'], "Concurrent simulated virtual user threads"),
                ("Test Duration (sec)", report['test_info']['actual_duration_sec'], "Total active execution time in seconds"),
                ("Total Requests Sent", report['throughput']['total_requests'], "Total HTTP requests dispatched to backend API"),
                ("Throughput (RPS)", report['throughput']['requests_per_second'], "Average requests completed per second"),
                ("Success Rate (%)", f"{report['throughput']['success_rate_pct']}%", f"{report['throughput']['successful_requests']} succeeded, {report['throughput']['failed_requests']} failed"),
                ("Average Latency (ms)", f"{report['response_time_ms']['avg']} ms", "Mean response latency across all requests"),
                ("p90 Latency (ms)", f"{report['response_time_ms']['p90']} ms", "90th percentile response time"),
                ("p95 Latency (ms)", f"{report['response_time_ms']['p95']} ms", "95th percentile response time"),
            ]

            for param, val, desc in summary_data:
                ws.append([param, val, desc])
                r = ws.max_row
                ws[f"A{r}"].font = bold_font
                ws[f"B{r}"].font = regular_font
                ws[f"C{r}"].font = regular_font
                for col in ["A", "B", "C"]:
                    ws[f"{col}{r}"].border = thin_border

            # Endpoint Breakdown Table
            ws.append([])
            ws.append([])
            ws.append(["Endpoint Breakdown", "", "", "", "", ""])
            ws.merge_cells(f"A{ws.max_row}:F{ws.max_row}")
            ws[f"A{ws.max_row}"].font = Font(name="Calibri", size=13, bold=True, color="1F4E78")

            ep_header_row = ws.max_row + 1
            ws.append(["Endpoint Name", "Requests", "RPS", "Min (ms)", "Avg (ms)", "Max (ms)"])
            for col in range(1, 7):
                cell = ws.cell(row=ep_header_row, column=col)
                cell.font = header_font
                cell.fill = navy_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for name, stats in report["endpoint_breakdown"].items():
                ws.append([name, stats['count'], stats['rps'], stats['min_ms'], stats['avg_ms'], stats['max_ms']])
                r = ws.max_row
                ws[f"A{r}"].font = bold_font
                for c_idx, col in enumerate(["A", "B", "C", "D", "E", "F"], start=1):
                    cell = ws[f"{col}{r}"]
                    cell.font = regular_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left")

            # Column auto-fit
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

            wb.save(filename)
        except Exception as e:
            print(f"[!] Warning: Could not generate openpyxl Excel report: {e}")

    def generate_markdown_report(self, report, filename):
        md_content = f"""# RejectionIQ - Baseline / Load Testing Report

**Execution Timestamp**: `{report['test_info']['timestamp']}`  
**Target Environment**: `{report['test_info']['target_host']}`  

---

## 📊 Executive Summary

| Parameter | Value |
|---|---|
| **Virtual Users (VUs)** | `{report['test_info']['virtual_users']}` concurrent users |
| **Test Duration** | `{report['test_info']['actual_duration_sec']}` seconds |
| **Total Requests Sent** | **{report['throughput']['total_requests']:,}** |
| **Requests Per Second (RPS)** | **{report['throughput']['requests_per_second']} req/sec** |
| **Success Rate** | **{report['throughput']['success_rate_pct']}%** ({report['throughput']['successful_requests']} succeeded, {report['throughput']['failed_requests']} failed) |

---

## ⏱️ Response Time Statistics

> [!NOTE]
> All response time metrics are measured end-to-end in milliseconds (ms).

| Metric | Latency (ms) | Target Benchmark | Description |
|---|---|---|---|
| **Fastest (Min)** | **{report['response_time_ms']['min']} ms** | `~50 ms` | Absolute fastest response time recorded |
| **Average (Avg)** | **{report['response_time_ms']['avg']} ms** | `~250 ms` | Mean response latency across all virtual users |
| **Median (p50)** | **{report['response_time_ms']['median_p50']} ms** | `~200 ms` | 50% of requests completed faster than this |
| **90th Percentile (p90)** | **{report['response_time_ms']['p90']} ms** | `~400 ms` | 90% of requests completed within this window |
| **95th Percentile (p95)** | **{report['response_time_ms']['p95']} ms** | `~450 ms` | 95% of requests completed within this window |
| **Slowest (Max)** | **{report['response_time_ms']['max']} ms** | `~1500 ms` | Maximum response time observed under heavy load |

---

## 🎯 Endpoint Throughput & Latency Breakdown

| Endpoint Name | Requests | RPS | Min (ms) | Avg (ms) | Max (ms) |
|---|---|---|---|---|---|
"""
        for name, stats in report["endpoint_breakdown"].items():
            md_content += f"| **{name}** | {stats['count']:,} | {stats['rps']} req/s | {stats['min_ms']} ms | {stats['avg_ms']} ms | {stats['max_ms']} ms |\n"

        md_content += f"""
---

## 🚦 HTTP Status Code Distribution

```json
{json.dumps(report['status_codes'], indent=2)}
```
"""
        with open(filename, "w", encoding="utf-8") as f:
            f.write(md_content)

def main():
    parser = argparse.ArgumentParser(description="RejectionIQ Baseline / Load Testing Suite")
    parser.add_argument("--host", type=str, default=DEFAULT_HOST, help="Target API base URL (default: http://127.0.0.1:8000)")
    parser.add_argument("--users", type=int, default=DEFAULT_USERS, help="Number of virtual concurrent users (default: 100)")
    parser.add_argument("--duration", type=int, default=DEFAULT_DURATION, help="Load test duration in seconds (default: 60)")

    args = parser.parse_args()
    tester = LoadTester(host=args.host, num_users=args.users, duration_sec=args.duration)
    tester.run()

if __name__ == "__main__":
    main()
