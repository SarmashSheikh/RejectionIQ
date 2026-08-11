import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ExcelReportGenerator:
    def __init__(self, test_results, output_dir):
        self.results = test_results
        self.output_dir = output_dir

    def generate_all_excel_reports(self):
        os.makedirs(self.output_dir, exist_ok=True)
        self._generate_main_report(os.path.join(self.output_dir, "Automation_Test_Report.xlsx"))
        self._generate_filtered_report(os.path.join(self.output_dir, "Passed_Test_Cases.xlsx"), "Passed")
        self._generate_filtered_report(os.path.join(self.output_dir, "Failed_Test_Cases.xlsx"), "Failed")
        self._generate_summary_report(os.path.join(self.output_dir, "Summary_Report.xlsx"))

    def _generate_main_report(self, filepath):
        wb = openpyxl.Workbook()
        
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Calibri", size=11, bold=True)
        font_regular = Font(name="Calibri", size=11)
        fill_navy = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        fill_blue = PatternFill(start_color="2C4D75", end_color="2C4D75", fill_type="solid")
        fill_pass = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        font_pass = Font(name="Calibri", size=11, bold=True, color="375623")
        fill_fail = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        font_fail = Font(name="Calibri", size=11, bold=True, color="C65911")
        border_cell = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

        # Sheet 1: Executed Test Cases
        ws_exec = wb.active
        ws_exec.title = "Executed Test Cases"
        ws_exec.views.sheetView[0].showGridLines = True
        
        headers = ["Test ID", "Module", "Test Name", "Status", "Execution Time (ms)", "Priority"]
        for c, h in enumerate(headers, 1):
            cell = ws_exec.cell(row=1, column=c, value=h)
            cell.font = font_header; cell.fill = fill_navy
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r_idx, r in enumerate(self.results, start=2):
            ws_exec.cell(row=r_idx, column=1, value=r["id"]).font = font_bold
            ws_exec.cell(row=r_idx, column=2, value=r["module"]).font = font_regular
            ws_exec.cell(row=r_idx, column=3, value=r["name"]).font = font_regular
            
            st_cell = ws_exec.cell(row=r_idx, column=4, value=r["status"])
            st_cell.alignment = Alignment(horizontal="center")
            if r["status"] == "Passed":
                st_cell.fill = fill_pass; st_cell.font = font_pass
            else:
                st_cell.fill = fill_fail; st_cell.font = font_fail

            ws_exec.cell(row=r_idx, column=5, value=r["duration_ms"]).alignment = Alignment(horizontal="center")
            ws_exec.cell(row=r_idx, column=6, value=r["priority"]).alignment = Alignment(horizontal="center")
            for col in range(1, 7): ws_exec.cell(row=r_idx, column=col).border = border_cell

        # Sheet 2: Passed Tests
        ws_pass = wb.create_sheet(title="Passed Tests")
        ws_pass.views.sheetView[0].showGridLines = True
        for c, h in enumerate(headers, 1):
            cell = ws_pass.cell(row=1, column=c, value=h); cell.font = font_header; cell.fill = fill_blue
        p_row = 2
        for r in self.results:
            if r["status"] == "Passed":
                ws_pass.cell(row=p_row, column=1, value=r["id"]).font = font_bold
                ws_pass.cell(row=p_row, column=2, value=r["module"])
                ws_pass.cell(row=p_row, column=3, value=r["name"])
                ws_pass.cell(row=p_row, column=4, value=r["status"]).fill = fill_pass
                ws_pass.cell(row=p_row, column=5, value=r["duration_ms"])
                ws_pass.cell(row=p_row, column=6, value=r["priority"])
                p_row += 1

        # Sheet 3: Failed Tests
        ws_fail = wb.create_sheet(title="Failed Tests")
        ws_fail.views.sheetView[0].showGridLines = True
        for c, h in enumerate(headers, 1):
            cell = ws_fail.cell(row=1, column=c, value=h); cell.font = font_header; cell.fill = fill_navy
        f_row = 2
        for r in self.results:
            if r["status"] == "Failed":
                ws_fail.cell(row=f_row, column=1, value=r["id"]).font = font_bold
                ws_fail.cell(row=f_row, column=2, value=r["module"])
                ws_fail.cell(row=f_row, column=3, value=r["name"])
                ws_fail.cell(row=f_row, column=4, value=r["status"]).fill = fill_fail
                ws_fail.cell(row=f_row, column=5, value=r["duration_ms"])
                ws_fail.cell(row=f_row, column=6, value=r["priority"])
                f_row += 1

        # Sheet 4: Skipped Tests
        ws_skip = wb.create_sheet(title="Skipped Tests")
        ws_skip.views.sheetView[0].showGridLines = True
        for c, h in enumerate(headers, 1):
            cell = ws_skip.cell(row=1, column=c, value=h); cell.font = font_header; cell.fill = fill_blue

        # Sheet 5: Execution Metrics
        ws_met = wb.create_sheet(title="Execution Metrics")
        ws_met["A1"] = "Execution Parameter"; ws_met["B1"] = "Value"
        ws_met["A1"].font = font_header; ws_met["A1"].fill = fill_navy
        ws_met["B1"].font = font_header; ws_met["B1"].fill = fill_navy
        tot = len(self.results)
        pas = sum(1 for r in self.results if r["status"] == "Passed")
        fai = sum(1 for r in self.results if r["status"] == "Failed")
        rate = (pas / tot * 100) if tot > 0 else 0
        metrics = [("Total Test Cases", tot), ("Passed Tests", pas), ("Failed Tests", fai), ("Success Rate %", f"{rate:.2f}%")]
        for m_idx, (k, v) in enumerate(metrics, start=2):
            ws_met.cell(row=m_idx, column=1, value=k).font = font_bold
            ws_met.cell(row=m_idx, column=2, value=v).alignment = Alignment(horizontal="center")

        # Sheet 6: Defect Summary
        ws_def = wb.create_sheet(title="Defect Summary")
        ws_def["A1"] = "Defect ID"; ws_def["B1"] = "Test ID"; ws_def["C1"] = "Module"; ws_def["D1"] = "Failure Reason"
        for c in range(1, 5): ws_def.cell(row=1, column=c).font = font_header; ws_def.cell(row=1, column=c).fill = fill_navy
        d_idx = 2
        for r in self.results:
            if r["status"] == "Failed":
                ws_def.cell(row=d_idx, column=1, value=f"DEF_{d_idx-1:03d}").font = font_bold
                ws_def.cell(row=d_idx, column=2, value=r["id"])
                ws_def.cell(row=d_idx, column=3, value=r["module"])
                ws_def.cell(row=d_idx, column=4, value=r["actual"])
                d_idx += 1

        wb.save(filepath)

    def _generate_filtered_report(self, filepath, target_status):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{target_status} Tests"
        headers = ["Test ID", "Module", "Test Name", "Status", "Execution Time (ms)", "Priority"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1B365D", fill_type="solid")
            
        r_row = 2
        for r in self.results:
            if r["status"] == target_status:
                ws.cell(row=r_row, column=1, value=r["id"])
                ws.cell(row=r_row, column=2, value=r["module"])
                ws.cell(row=r_row, column=3, value=r["name"])
                ws.cell(row=r_row, column=4, value=r["status"])
                ws.cell(row=r_row, column=5, value=r["duration_ms"])
                ws.cell(row=r_row, column=6, value=r["priority"])
                r_row += 1
        wb.save(filepath)

    def _generate_summary_report(self, filepath):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary Report"
        ws["A1"] = "Metric"; ws["B1"] = "Value"
        tot = len(self.results)
        pas = sum(1 for r in self.results if r["status"] == "Passed")
        fai = sum(1 for r in self.results if r["status"] == "Failed")
        rate = (pas / tot * 100) if tot > 0 else 0
        ws.append(["Total Test Cases", tot])
        ws.append(["Passed Test Cases", pas])
        ws.append(["Failed Test Cases", fai])
        ws.append(["Pass Percentage", f"{rate:.2f}%"])
        wb.save(filepath)
